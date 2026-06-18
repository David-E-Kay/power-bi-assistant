#!/usr/bin/env python3
"""
compare-snapshots.py — Unified regression test & timing comparison
===================================================================
Compares two JSON snapshots produced by capture-snapshot.csx and outputs
a single formatted .xlsx covering both value parity and performance timing.

Dependencies: openpyxl (declared in requirements.txt)

Usage:
    python compare-snapshots.py baseline.json refactored.json [--output regression-report.xlsx]

Exit codes:
    0 = all value tests pass (timing regressions don't affect exit code)
    1 = one or more value failures detected

Sheets:
    All Tests          — every test case: value delta flag, timing side-by-side
    Value Deltas       — cell-level mismatch detail (only where Delta=Y)
    By Measure         — timing aggregated per measure
    By Context         — timing aggregated per context
    Top Movers         — top 20 timing regressions + improvements
    Timeout Regressions — tests that newly timed out vs baseline (NEW)
"""

import argparse
import json
import os
import sys
import csv
from collections import defaultdict, OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required for the Excel report. Install it with:\n"
             "    pip install -r requirements.txt\n"
             "(or directly: pip install openpyxl)")

# ═════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═════════════════════════════════════════════════════════════════════════════
NUMERIC_TOLERANCE = 1e-4       # value comparison tolerance
REGRESSION_PCT = 20            # timing: ≥ +20% slower → regression
IMPROVEMENT_PCT = -20          # timing: ≤ -20% faster → improvement
MIN_MS_FOR_PCT = 50            # timing: ignore % on tests under 50ms

# ═════════════════════════════════════════════════════════════════════════════
# STYLES
# ═════════════════════════════════════════════════════════════════════════════
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

FONT = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)

DELTA_Y_FILL = PatternFill("solid", fgColor="FFC7CE")
DELTA_Y_FONT = Font(name="Arial", size=10, color="9C0006", bold=True)
DELTA_N_FILL = PatternFill("solid", fgColor="C6EFCE")
DELTA_N_FONT = Font(name="Arial", size=10, color="006100")

REG_FILL = PatternFill("solid", fgColor="FFC7CE")
REG_FONT = Font(name="Arial", size=10, color="9C0006")
IMP_FILL = PatternFill("solid", fgColor="C6EFCE")
IMP_FONT = Font(name="Arial", size=10, color="006100")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# ═════════════════════════════════════════════════════════════════════════════
# DATA LOADING & VALUE COMPARISON
# ═════════════════════════════════════════════════════════════════════════════

def load_snapshot(path):
    with open(path, "r", encoding="utf-8-sig") as f:
        return json.load(f)


def values_equal(a, b):
    if a == b:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) < NUMERIC_TOLERANCE
    return False


def identify_groupby_cols(columns):
    return [c for c in columns if c not in ("[Result]", "Result")]


def compare_test_case(tid, base_tc, refac_tc):
    """Compare values for a single test case.
    Returns (delta_flag, status, detail_text, cell_mismatches).
    """
    measure = base_tc.get("measure", tid)
    context = base_tc.get("context", "unknown")

    # New statuses — timeout
    b_timeout = base_tc.get("status") == "timeout"
    r_timeout = refac_tc.get("status") == "timeout"

    # New statuses — skipped / aborted
    b_skipped = base_tc.get("status") == "skipped"
    r_skipped = refac_tc.get("status") == "skipped"
    b_aborted = base_tc.get("status") == "aborted_memory"
    r_aborted = refac_tc.get("status") == "aborted_memory"

    # Error handling
    b_err = base_tc.get("status") == "error"
    r_err = refac_tc.get("status") == "error"

    # Handle all non-ok statuses
    if b_timeout and r_timeout:
        return "N", "timeout_both", "Both timed out", []
    if b_timeout:
        return "Y", "timeout_base", f"Baseline timeout ({base_tc.get('timeout_limit_ms', '')}ms limit)", []
    if r_timeout:
        return "Y", "timeout_refac", f"Refactored timeout ({refac_tc.get('timeout_limit_ms', '')}ms limit)", []
    if b_aborted and r_aborted:
        return "N", "aborted_both", "Both aborted (memory watchdog)", []
    if b_aborted:
        return "Y", "aborted_base", "Baseline aborted by memory watchdog", []
    if r_aborted:
        return "Y", "aborted_refac", "Refactored aborted by memory watchdog", []
    if b_skipped and r_skipped:
        return "N", "skipped_both", "Both skipped/aborted", []
    if b_skipped:
        return "Y", "skipped_base", f"Baseline skipped: {base_tc.get('skip_reason', base_tc.get('status', ''))[:60]}", []
    if r_skipped:
        return "Y", "skipped_refac", f"Refactored skipped: {refac_tc.get('skip_reason', refac_tc.get('status', ''))[:60]}", []
    if b_err and r_err:
        return "Y", "error_both", "Both errored", []
    if b_err:
        return "Y", "error_base", f"Baseline error: {base_tc.get('error', '')[:80]}", []
    if r_err:
        return "Y", "error_refac", f"Refactored error: {refac_tc.get('error', '')[:80]}", []

    base_rows = base_tc.get("rows", [])
    refac_rows = refac_tc.get("rows", [])
    base_cols = base_tc.get("columns", [])

    # Row count mismatch
    if len(base_rows) != len(refac_rows):
        detail = f"Row count: {len(base_rows)} → {len(refac_rows)}"
        return "Y", "row_count", detail, [{
            "test_id": tid, "measure": measure, "context": context,
            "row_key": "ROW_COUNT", "column": "",
            "baseline": len(base_rows), "refactored": len(refac_rows),
        }]

    # Both empty
    if len(base_rows) == 0:
        return "N", "pass", "", []

    # Sort by groupby columns
    groupby = identify_groupby_cols(base_cols)
    if groupby:
        base_rows = sorted(base_rows, key=lambda r: tuple(str(r.get(c, "")) for c in groupby))
        refac_rows = sorted(refac_rows, key=lambda r: tuple(str(r.get(c, "")) for c in groupby))

    # Cell-by-cell comparison
    mismatches = []
    for i, (br, rr) in enumerate(zip(base_rows, refac_rows)):
        for col in base_cols:
            bv = br.get(col)
            rv = rr.get(col)
            if not values_equal(bv, rv):
                row_key = " | ".join(f"{c}={br.get(c)}" for c in groupby) if groupby else f"row {i}"
                mismatches.append({
                    "test_id": tid, "measure": measure, "context": context,
                    "row_key": row_key, "column": col,
                    "baseline": bv, "refactored": rv,
                })

    if mismatches:
        detail = f"{len(mismatches)} cell(s) differ"
        return "Y", "fail", detail, mismatches[:20]  # cap detail rows per test

    return "N", "pass", "", []


# ═════════════════════════════════════════════════════════════════════════════
# TIMING COMPARISON
# ═════════════════════════════════════════════════════════════════════════════

def classify_timing(baseline_ms, refactored_ms):
    delta = refactored_ms - baseline_ms
    if baseline_ms >= MIN_MS_FOR_PCT and baseline_ms > 0:
        pct = (delta / baseline_ms) * 100
    else:
        pct = 0.0 if delta == 0 else (100.0 if delta > 0 else -100.0)

    if pct >= REGRESSION_PCT and delta > MIN_MS_FOR_PCT:
        verdict = "REGRESSION"
    elif pct <= IMPROVEMENT_PCT and abs(delta) > MIN_MS_FOR_PCT:
        verdict = "IMPROVEMENT"
    else:
        verdict = ""
    return delta, round(pct, 1), verdict


# ═════════════════════════════════════════════════════════════════════════════
# XLSX HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def style_cell(cell, bold=False):
    cell.font = FONT_BOLD if bold else FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")


def auto_width(ws, min_w=10, max_w=45):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        widths = [len(str(c.value or "")) for c in col_cells]
        ws.column_dimensions[letter].width = min(max(max(widths) + 2, min_w), max_w) if widths else min_w


# ═════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═════════════════════════════════════════════════════════════════════════════

def build_all_tests_sheet(wb, joined, base_meta, refac_meta):
    ws = wb.active
    ws.title = "All Tests"

    # Meta header
    ws.merge_cells("A1:L1")
    ws["A1"].value = "Regression Test Report"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="2F5496")

    ws["A2"] = "Baseline:"
    ws["B2"] = base_meta.get("label", "")
    ws["C2"] = base_meta.get("captured_at", "")[:19]
    ws["A3"] = "Refactored:"
    ws["B3"] = refac_meta.get("label", "")
    ws["C3"] = refac_meta.get("captured_at", "")[:19]
    for r in (2, 3):
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=2).font = FONT
        ws.cell(row=r, column=3).font = FONT

    ws["A4"] = (f"Timing thresholds: regression >= +{REGRESSION_PCT}% and +{MIN_MS_FOR_PCT}ms, "
                f"improvement <= {IMPROVEMENT_PCT}% and -{MIN_MS_FOR_PCT}ms  |  "
                f"Value tolerance: {NUMERIC_TOLERANCE}")
    ws["A4"].font = Font(name="Arial", size=9, italic=True, color="666666")

    # Column headers
    headers = [
        "Test ID", "Measure", "Context",
        "Delta", "Delta Detail",
        "Baseline (ms)", "Refactored (ms)", "Δ (ms)", "Δ (%)", "Timing Verdict",
        "Baseline Rows", "Refactored Rows",
    ]
    hrow = 6
    for col, h in enumerate(headers, 1):
        ws.cell(row=hrow, column=col, value=h)
    style_header(ws, hrow, len(headers))

    # Data rows
    row = hrow + 1
    for e in joined:
        ws.cell(row=row, column=1, value=e["test_id"])
        ws.cell(row=row, column=2, value=e["measure"])
        ws.cell(row=row, column=3, value=e["context"])

        delta_cell = ws.cell(row=row, column=4, value=e["delta_flag"])
        ws.cell(row=row, column=5, value=e["delta_detail"])
        ws.cell(row=row, column=6, value=e["baseline_ms"])
        ws.cell(row=row, column=7, value=e["refactored_ms"])
        ws.cell(row=row, column=8, value=e["timing_delta_ms"])

        pct_cell = ws.cell(row=row, column=9,
                           value=e["timing_delta_pct"] / 100 if e["timing_delta_pct"] else 0)
        pct_cell.number_format = "0.0%"

        ws.cell(row=row, column=10, value=e["timing_verdict"])
        ws.cell(row=row, column=11, value=e["baseline_rows"])
        ws.cell(row=row, column=12, value=e["refactored_rows"])

        for col in range(1, len(headers) + 1):
            style_cell(ws.cell(row=row, column=col))

        # Delta flag styling
        if e["delta_flag"] == "Y":
            delta_cell.fill = DELTA_Y_FILL
            delta_cell.font = DELTA_Y_FONT
        else:
            delta_cell.fill = DELTA_N_FILL
            delta_cell.font = DELTA_N_FONT

        # Timing verdict styling
        v = e["timing_verdict"]
        if v == "REGRESSION":
            for c in (8, 9, 10):
                ws.cell(row=row, column=c).fill = REG_FILL
                ws.cell(row=row, column=c).font = REG_FONT
        elif v == "IMPROVEMENT":
            for c in (8, 9, 10):
                ws.cell(row=row, column=c).fill = IMP_FILL
                ws.cell(row=row, column=c).font = IMP_FONT

        row += 1

    # Summary formulas
    sr = row + 1
    ws.cell(row=sr, column=1, value="TOTALS").font = FONT_BOLD
    for c in (6, 7, 8):
        ws.cell(row=sr, column=c, value=f"=SUM({get_column_letter(c)}{hrow+1}:{get_column_letter(c)}{row-1})")
        ws.cell(row=sr, column=c).font = FONT_BOLD
        ws.cell(row=sr, column=c).border = THIN_BORDER

    # Value delta summary
    sr2 = sr + 1
    ws.cell(row=sr2, column=1, value="Value Deltas").font = FONT_BOLD
    ws.cell(row=sr2, column=4, value=f'=COUNTIF(D{hrow+1}:D{row-1},"Y")')
    ws.cell(row=sr2, column=4).font = FONT_BOLD
    ws.cell(row=sr2, column=4).border = THIN_BORDER
    ws.cell(row=sr2, column=5, value=f"of {row - hrow - 1}").font = FONT

    ws.freeze_panes = f"A{hrow + 1}"
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(headers))}{row - 1}"
    auto_width(ws)


def build_value_deltas_sheet(wb, all_mismatches):
    ws = wb.create_sheet("Value Deltas")

    headers = ["Test ID", "Measure", "Context", "Row Key", "Column",
               "Baseline Value", "Refactored Value"]

    if not all_mismatches:
        ws["A1"] = "No value deltas detected — all tests passed."
        ws["A1"].font = Font(name="Arial", size=12, bold=True, color="006100")
        return

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for mm in all_mismatches:
        ws.cell(row=row, column=1, value=mm["test_id"])
        ws.cell(row=row, column=2, value=mm["measure"])
        ws.cell(row=row, column=3, value=mm["context"])
        ws.cell(row=row, column=4, value=mm["row_key"])
        ws.cell(row=row, column=5, value=mm["column"])
        ws.cell(row=row, column=6, value=str(mm["baseline"]))
        ws.cell(row=row, column=7, value=str(mm["refactored"]))

        for col in range(1, len(headers) + 1):
            style_cell(ws.cell(row=row, column=col))
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    auto_width(ws)


def build_aggregation_sheet(ws, joined, group_key, group_label):
    """Shared builder for By Measure and By Context sheets."""
    agg = defaultdict(lambda: {"b_total": 0, "r_total": 0, "count": 0,
                                "regs": 0, "imps": 0, "deltas": 0})
    for e in joined:
        g = e[group_key]
        agg[g]["b_total"] += e["baseline_ms"]
        agg[g]["r_total"] += e["refactored_ms"]
        agg[g]["count"] += 1
        if e["timing_verdict"] == "REGRESSION":
            agg[g]["regs"] += 1
        elif e["timing_verdict"] == "IMPROVEMENT":
            agg[g]["imps"] += 1
        if e["delta_flag"] == "Y":
            agg[g]["deltas"] += 1

    headers = [
        group_label, "Tests", "Value Deltas",
        "Baseline Total (ms)", "Refactored Total (ms)", "Δ Total (ms)",
        "Avg Baseline (ms)", "Avg Refactored (ms)", "Δ Avg (ms)",
        "Timing Regressions", "Timing Improvements",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    sorted_groups = sorted(agg.items(), key=lambda x: x[1]["r_total"] - x[1]["b_total"], reverse=True)

    row = 2
    for name, s in sorted_groups:
        delta = s["r_total"] - s["b_total"]
        avg_b = round(s["b_total"] / s["count"]) if s["count"] else 0
        avg_r = round(s["r_total"] / s["count"]) if s["count"] else 0

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=s["count"])

        d_cell = ws.cell(row=row, column=3, value=s["deltas"])
        if s["deltas"] > 0:
            d_cell.fill = DELTA_Y_FILL
            d_cell.font = DELTA_Y_FONT

        ws.cell(row=row, column=4, value=s["b_total"])
        ws.cell(row=row, column=5, value=s["r_total"])
        ws.cell(row=row, column=6, value=delta)
        ws.cell(row=row, column=7, value=avg_b)
        ws.cell(row=row, column=8, value=avg_r)
        ws.cell(row=row, column=9, value=avg_r - avg_b)

        reg_cell = ws.cell(row=row, column=10, value=s["regs"])
        imp_cell = ws.cell(row=row, column=11, value=s["imps"])

        for col in range(1, len(headers) + 1):
            style_cell(ws.cell(row=row, column=col))

        if s["regs"] > 0:
            reg_cell.fill = REG_FILL
            reg_cell.font = REG_FONT
        if s["imps"] > 0:
            imp_cell.fill = IMP_FILL
            imp_cell.font = IMP_FONT

        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    auto_width(ws)


def build_top_movers_sheet(wb, joined, top_n=20):
    ws = wb.create_sheet("Top Movers")

    headers = ["Test ID", "Measure", "Context", "Delta",
               "Baseline (ms)", "Refactored (ms)", "Δ (ms)", "Δ (%)", "Timing Verdict"]

    # Top regressions
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Top {top_n} Timing Regressions (slowest Δ)"
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color="9C0006")

    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header(ws, 2, len(headers))

    slowest = sorted(joined, key=lambda x: x["timing_delta_ms"], reverse=True)[:top_n]
    row = 3
    for e in slowest:
        ws.cell(row=row, column=1, value=e["test_id"])
        ws.cell(row=row, column=2, value=e["measure"])
        ws.cell(row=row, column=3, value=e["context"])
        d_cell = ws.cell(row=row, column=4, value=e["delta_flag"])
        ws.cell(row=row, column=5, value=e["baseline_ms"])
        ws.cell(row=row, column=6, value=e["refactored_ms"])
        ws.cell(row=row, column=7, value=e["timing_delta_ms"])
        pct = ws.cell(row=row, column=8, value=e["timing_delta_pct"] / 100 if e["timing_delta_pct"] else 0)
        pct.number_format = "0.0%"
        ws.cell(row=row, column=9, value=e["timing_verdict"])
        for col in range(1, len(headers) + 1):
            style_cell(ws.cell(row=row, column=col))
        if e["delta_flag"] == "Y":
            d_cell.fill = DELTA_Y_FILL
            d_cell.font = DELTA_Y_FONT
        if e["timing_verdict"] == "REGRESSION":
            for c in (7, 8, 9):
                ws.cell(row=row, column=c).fill = REG_FILL
                ws.cell(row=row, column=c).font = REG_FONT
        row += 1

    # Top improvements
    gap = row + 1
    ws.merge_cells(f"A{gap}:I{gap}")
    ws[f"A{gap}"] = f"Top {top_n} Timing Improvements (fastest Δ)"
    ws[f"A{gap}"].font = Font(name="Arial", size=12, bold=True, color="006100")

    hrow2 = gap + 1
    for col, h in enumerate(headers, 1):
        ws.cell(row=hrow2, column=col, value=h)
    style_header(ws, hrow2, len(headers))

    fastest = sorted(joined, key=lambda x: x["timing_delta_ms"])[:top_n]
    row = hrow2 + 1
    for e in fastest:
        ws.cell(row=row, column=1, value=e["test_id"])
        ws.cell(row=row, column=2, value=e["measure"])
        ws.cell(row=row, column=3, value=e["context"])
        d_cell = ws.cell(row=row, column=4, value=e["delta_flag"])
        ws.cell(row=row, column=5, value=e["baseline_ms"])
        ws.cell(row=row, column=6, value=e["refactored_ms"])
        ws.cell(row=row, column=7, value=e["timing_delta_ms"])
        pct = ws.cell(row=row, column=8, value=e["timing_delta_pct"] / 100 if e["timing_delta_pct"] else 0)
        pct.number_format = "0.0%"
        ws.cell(row=row, column=9, value=e["timing_verdict"])
        for col in range(1, len(headers) + 1):
            style_cell(ws.cell(row=row, column=col))
        if e["delta_flag"] == "Y":
            d_cell.fill = DELTA_Y_FILL
            d_cell.font = DELTA_Y_FONT
        if e["timing_verdict"] == "IMPROVEMENT":
            for c in (7, 8, 9):
                ws.cell(row=row, column=c).fill = IMP_FILL
                ws.cell(row=row, column=c).font = IMP_FONT
        row += 1

    auto_width(ws)


def build_timeout_regressions_sheet(wb, joined):
    ws = wb.create_sheet("Timeout Regressions")

    timeout_regressions = [e for e in joined
                           if e.get("delta_status") in ("timeout_refac", "timeout_base")]

    if not timeout_regressions:
        ws["A1"] = "No timeout regressions — all tests completed within the time limit."
        ws["A1"].font = Font(name="Arial", size=12, bold=True, color="006100")
        return

    headers = ["Test ID", "Measure", "Context", "Direction",
               "Baseline (ms)", "Refactored (ms)", "Timeout Limit (ms)", "DAX (truncated)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for e in timeout_regressions:
        direction = "New timeout (regression)" if e["delta_status"] == "timeout_refac" else "Fixed timeout (improvement)"
        ws.cell(row=row, column=1, value=e["test_id"])
        ws.cell(row=row, column=2, value=e["measure"])
        ws.cell(row=row, column=3, value=e["context"])
        dir_cell = ws.cell(row=row, column=4, value=direction)
        ws.cell(row=row, column=5, value=e["baseline_ms"])
        ws.cell(row=row, column=6, value=e["refactored_ms"])
        # timeout_limit_ms already resolved in main loop from whichever snapshot had the timeout entry
        ws.cell(row=row, column=7, value=e.get("timeout_limit_ms", ""))
        ws.cell(row=row, column=8, value=e.get("dax_truncated", ""))
        for col in range(1, len(headers) + 1):
            style_cell(ws.cell(row=row, column=col))
        if e["delta_status"] == "timeout_refac":
            dir_cell.fill = REG_FILL
            dir_cell.font = REG_FONT
        else:
            dir_cell.fill = IMP_FILL
            dir_cell.font = IMP_FONT
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    auto_width(ws)


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Compare regression test snapshots — values + timing in one report."
    )
    parser.add_argument("baseline", help="Path to baseline snapshot JSON")
    parser.add_argument("refactored", help="Path to refactored snapshot JSON")
    parser.add_argument("--output", "-o", default="regression-report.xlsx",
                        help="Output xlsx path (default: regression-report.xlsx)")
    args = parser.parse_args()

    base = load_snapshot(args.baseline)
    refac = load_snapshot(args.refactored)

    base_results = base.get("results", {})
    refac_results = refac.get("results", {})
    all_ids = sorted(set(base_results.keys()) | set(refac_results.keys()))

    joined = []
    all_mismatches = []
    value_counts = {"pass": 0, "fail": 0, "row_count": 0,
                    "error_both": 0, "error_base": 0, "error_refac": 0, "missing": 0,
                    "timeout_both": 0, "timeout_base": 0, "timeout_refac": 0,
                    "skipped_both": 0, "skipped_base": 0, "skipped_refac": 0,
                    "aborted_both": 0, "aborted_base": 0, "aborted_refac": 0}
    timing_counts = {"regression": 0, "improvement": 0, "neutral": 0}

    for tid in all_ids:
        b = base_results.get(tid)
        r = refac_results.get(tid)

        # ── Missing in one side ───────────────────────────────────────────
        if not b or not r:
            value_counts["missing"] += 1
            src = b or r
            joined.append({
                "test_id": tid,
                "measure": src.get("measure", ""),
                "context": src.get("context", ""),
                "delta_flag": "Y",
                "delta_detail": "missing in " + ("baseline" if not b else "refactored"),
                "delta_status": "missing",
                "baseline_ms": b.get("duration_ms", 0) if b else 0,
                "refactored_ms": r.get("duration_ms", 0) if r else 0,
                "timing_delta_ms": 0, "timing_delta_pct": 0, "timing_verdict": "",
                "baseline_rows": b.get("row_count", 0) if b else 0,
                "refactored_rows": r.get("row_count", 0) if r else 0,
                "timeout_limit_ms": "",
                "dax_truncated": "",
            })
            continue

        # ── Value comparison ──────────────────────────────────────────────
        delta_flag, status, detail, mismatches = compare_test_case(tid, b, r)
        value_counts[status] = value_counts.get(status, 0) + 1
        all_mismatches.extend(mismatches)

        # ── Timing comparison ─────────────────────────────────────────────
        b_ms = b.get("duration_ms", 0)
        r_ms = r.get("duration_ms", 0)
        t_delta, t_pct, t_verdict = classify_timing(b_ms, r_ms)

        if t_verdict == "REGRESSION":
            timing_counts["regression"] += 1
        elif t_verdict == "IMPROVEMENT":
            timing_counts["improvement"] += 1
        else:
            timing_counts["neutral"] += 1

        # Determine timeout_limit_ms and dax from whichever side has the timeout entry
        timeout_limit_ms = ""
        dax_truncated = ""
        if r and r.get("status") == "timeout":
            timeout_limit_ms = r.get("timeout_limit_ms", "")
            dax_truncated = (r.get("dax", "") or "")[:200]
        elif b and b.get("status") == "timeout":
            timeout_limit_ms = b.get("timeout_limit_ms", "")
            dax_truncated = (b.get("dax", "") or "")[:200]

        joined.append({
            "test_id": tid,
            "measure": b.get("measure", ""),
            "context": b.get("context", ""),
            "delta_flag": delta_flag,
            "delta_detail": detail,
            "delta_status": status,
            "baseline_ms": b_ms,
            "refactored_ms": r_ms,
            "timing_delta_ms": t_delta,
            "timing_delta_pct": t_pct,
            "timing_verdict": t_verdict,
            "baseline_rows": b.get("row_count", 0),
            "refactored_rows": r.get("row_count", 0),
            "timeout_limit_ms": timeout_limit_ms,
            "dax_truncated": dax_truncated,
        })

    # Sort: value deltas first, then by timing delta descending
    joined.sort(key=lambda x: (0 if x["delta_flag"] == "Y" else 1, -x["timing_delta_ms"]))

    # ── Build workbook ────────────────────────────────────────────────────
    wb = Workbook()
    build_all_tests_sheet(wb, joined, base, refac)
    build_value_deltas_sheet(wb, all_mismatches)

    ws_measure = wb.create_sheet("By Measure")
    build_aggregation_sheet(ws_measure, joined, "measure", "Measure")

    ws_context = wb.create_sheet("By Context")
    build_aggregation_sheet(ws_context, joined, "context", "Context")

    build_top_movers_sheet(wb, joined)
    build_timeout_regressions_sheet(wb, joined)
    wb.save(args.output)

    # ── Console summary ───────────────────────────────────────────────────
    total = len(all_ids)
    delta_count = sum(1 for e in joined if e["delta_flag"] == "Y")
    total_b = sum(e["baseline_ms"] for e in joined)
    total_r = sum(e["refactored_ms"] for e in joined)
    total_d = total_r - total_b

    print()
    print("━" * 62)
    print("  Regression Test Report")
    print(f"  Model: {base.get('model_name', 'Unknown')}")
    print(f"  Baseline:   {base.get('captured_at', '?')[:19]}")
    print(f"  Refactored: {refac.get('captured_at', '?')[:19]}")
    print("━" * 62)

    print()
    print("  VALUE COMPARISON")
    print(f"    ✅ Pass:              {value_counts['pass']:>4} / {total}")
    print(f"    ❌ Fail:              {value_counts['fail']:>4} / {total}")
    print(f"    🔢 Row count:         {value_counts['row_count']:>4} / {total}")
    print(f"    🔲 Missing:           {value_counts['missing']:>4} / {total}")
    errs = value_counts['error_both'] + value_counts['error_base'] + value_counts['error_refac']
    print(f"    ⚠️  Errors:            {errs:>4} / {total}")
    timeouts = value_counts['timeout_both'] + value_counts['timeout_base'] + value_counts['timeout_refac']
    skipped = (value_counts['skipped_both'] + value_counts['skipped_base'] + value_counts['skipped_refac']
               + value_counts['aborted_both'] + value_counts['aborted_base'] + value_counts['aborted_refac'])
    print(f"    ⏱️  Timeouts:           {timeouts:>4} / {total}")
    print(f"    ⏭️  Skipped/aborted:    {skipped:>4} / {total}")
    print(f"    ── Delta = Y:         {delta_count:>4} / {total}")

    print()
    print("  TIMING COMPARISON")
    print(f"    Baseline total:       {total_b:>8,} ms ({total_b/1000:.1f}s)")
    print(f"    Refactored total:     {total_r:>8,} ms ({total_r/1000:.1f}s)")
    print(f"    Δ total:              {total_d:>+8,} ms ({total_d/1000:+.1f}s)")
    if total_b > 0:
        print(f"    Δ overall:            {(total_d/total_b)*100:>+7.1f}%")
    print(f"    Regressions:          {timing_counts['regression']:>4}")
    print(f"    Improvements:         {timing_counts['improvement']:>4}")

    print()
    print(f"  Output: {os.path.abspath(args.output)}")
    print("━" * 62)

    # Exit code based on value failures only
    has_failures = (value_counts["fail"] + value_counts["row_count"]
                    + value_counts["missing"] + value_counts["error_base"]
                    + value_counts["error_refac"]
                    + value_counts["timeout_refac"]
                    + value_counts["aborted_refac"]) > 0

    # Desktop toast (best-effort; keeps the comparator runnable standalone even
    # if the pbi_capture package is not importable). Click opens the xlsx.
    try:
        from pathlib import Path
        from pbi_capture.notify import send_desktop_toast
        clean = delta_count == 0 and not has_failures
        status = "✅ No deltas" if clean else (
            f"⚠️ {delta_count} deltas / {timing_counts['regression']} regressions")
        body = [f"Report: {os.path.basename(args.output)}",
                f"Pass {value_counts['pass']} / Fail {value_counts['fail']} of {total}",
                f"Delta=Y: {delta_count}"]
        if total_b > 0:
            body.append(f"Δ overall: {(total_d / total_b) * 100:+.1f}%")
        send_desktop_toast(f"PBI Regression Compare — {status}", body,
                           launch_uri=Path(args.output).resolve().as_uri())
    except Exception:
        pass

    return 1 if has_failures else 0


if __name__ == "__main__":
    sys.exit(main())
