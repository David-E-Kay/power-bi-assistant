"""Benchmark Excel report builder.

Renders a polished multi-sheet .xlsx from a single benchmark run's timing rows,
mirroring the regression report (compare-snapshots.py) styling so the two
deliverables look like a set. openpyxl is imported lazily inside the writer, so
importing this module is cheap and the dependency stays optional (run_benchmark
treats a missing openpyxl as non-fatal). The data-prep helpers are pure and
openpyxl-free for unit testing.

Sheets: Summary | All Tests | By Measure | Slowest | False-Fast Warnings
"""
from collections import OrderedDict

SLOWEST_TOP_N = 20


# ── pure data-prep (no openpyxl) ─────────────────────────────────────────────

def false_fast(timing_rows):
    """distinct_values==1 with row_count>1 on a non-grand-total context — the
    dimension isn't filtering the measure (no relationship path). Same predicate
    as runner._false_fast; kept here so the report is self-contained."""
    return [t for t in timing_rows
            if t["status"] == "ok" and t["context"] != "grand_total"
            and t.get("distinct_values") == 1 and t["row_count"] > 1]


def slowest(timing_rows, top_n=SLOWEST_TOP_N):
    ok = [t for t in timing_rows if t["status"] == "ok"]
    return sorted(ok, key=lambda t: -t["duration_ms"])[:top_n]


def aggregate_by_measure(timing_rows):
    """Per-measure timing rollup over ok rows, sorted by total time descending."""
    agg = OrderedDict()
    for t in timing_rows:
        a = agg.setdefault(t["measure"], {
            "measure": t["measure"], "tests": 0, "ok": 0,
            "total_ms": 0, "max_ms": 0, "min_ms": None})
        a["tests"] += 1
        if t["status"] == "ok":
            a["ok"] += 1
            d = t["duration_ms"]
            a["total_ms"] += d
            a["max_ms"] = max(a["max_ms"], d)
            a["min_ms"] = d if a["min_ms"] is None else min(a["min_ms"], d)
    out = list(agg.values())
    for a in out:
        a["avg_ms"] = round(a["total_ms"] / a["ok"]) if a["ok"] else 0
        a["min_ms"] = a["min_ms"] or 0
    out.sort(key=lambda a: a["total_ms"], reverse=True)
    return out


# ── xlsx writer (openpyxl imported lazily) ───────────────────────────────────

def write_benchmark_report(path, *, cfg, counts, total, total_ms, timing_rows,
                           skipped, smoke_results, memory_aborted):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    FONT = Font(name="Arial", size=10)
    FONT_BOLD = Font(name="Arial", size=10, bold=True)
    TITLE_FONT = Font(name="Arial", size=14, bold=True, color="2F5496")
    BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
    BAD_FONT = Font(name="Arial", size=10, color="9C0006", bold=True)
    WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
    WARN_FONT = Font(name="Arial", size=10, color="9C6500", bold=True)
    OK_FONT = Font(name="Arial", size=12, bold=True, color="006100")
    BORDER = Border(*(Side(style="thin", color="B4C6E7"),) * 4)

    def header(ws, row, n):
        for col in range(1, n + 1):
            c = ws.cell(row=row, column=col)
            c.font, c.fill, c.alignment, c.border = (
                HEADER_FONT, HEADER_FILL, HEADER_ALIGN, BORDER)

    def cell(ws, r, c, value, bold=False):
        x = ws.cell(row=r, column=c, value=value)
        x.font = FONT_BOLD if bold else FONT
        x.border = BORDER
        x.alignment = Alignment(vertical="center")
        return x

    def auto_width(ws, min_w=10, max_w=50):
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            widths = [len(str(c.value or "")) for c in col_cells]
            ws.column_dimensions[letter].width = (
                min(max(max(widths) + 2, min_w), max_w) if widths else min_w)

    flagged = false_fast(timing_rows)
    flagged_ids = {t["test_id"] for t in flagged}
    aborted = counts.get("aborted_memory", 0)

    wb = Workbook()

    # ── Summary ──────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:B1")
    ws["A1"].value = "Measure Benchmark Report"
    ws["A1"].font = TITLE_FONT

    contexts = (f"1 grand_total + {len(cfg.single_slice_dimensions)} single-slice"
                + (" + 1 cross-product" if cfg.cross_product_columns else ""))
    facts = [
        ("Label", cfg.label),
        ("Measures", len(cfg.measures)),
        ("Contexts", contexts),
        ("Test cases", total),
        ("OK", counts["ok"]),
        ("Errors", counts["error"]),
        ("Timeouts", counts["timeout"]),
        ("Skipped", counts["skipped"]),
    ]
    if aborted:
        facts.append(("Aborted (memory)", aborted))
    facts += [
        ("Duration", f"{total_ms / 60000:.1f} min"),
        ("Query timeout", f"{cfg.query_timeout_ms} ms (ADOMD direct)"),
        ("Global filters", f"{len(cfg.global_filters)} (TREATAS)"),
        ("Cross-product columns", len(cfg.cross_product_columns)),
        ("False-fast warnings", len(flagged)),
    ]
    if cfg.max_rows_per_context > 0:
        facts.append(("Row cap", f"TOPN({cfg.max_rows_per_context}) per context"))
    if skipped:
        facts.append(("Smoke-test skipped", f"{len(skipped)} measure(s)"))
    if memory_aborted:
        facts.append(("Memory abort", f"threshold {cfg.memory_threshold_pct}%"))

    r = 3
    for k, v in facts:
        cell(ws, r, 1, k, bold=True)
        vc = cell(ws, r, 2, v)
        if k in ("Errors", "Timeouts") and v:
            vc.fill, vc.font = BAD_FILL, BAD_FONT
        elif k == "False-fast warnings" and v:
            vc.fill, vc.font = WARN_FILL, WARN_FONT
        r += 1
    for m, reason in smoke_results.items():
        cell(ws, r, 1, "  skipped", bold=True)
        cell(ws, r, 2, f"{m}: {reason}")
        r += 1
    auto_width(ws)

    # ── All Tests ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("All Tests")
    cols = ["Test ID", "Measure", "Context", "Status", "Rows",
            "Duration (ms)", "Distinct", "Flag"]
    for c, h in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=h)
    header(ws, 1, len(cols))
    row = 2
    for t in timing_rows:
        ff = t["test_id"] in flagged_ids
        flag = "FALSE-FAST" if ff else ("" if t["status"] == "ok" else t["status"].upper())
        cell(ws, row, 1, t["test_id"])
        cell(ws, row, 2, t["measure"])
        cell(ws, row, 3, t["context"])
        st = cell(ws, row, 4, t["status"])
        cell(ws, row, 5, t["row_count"])
        cell(ws, row, 6, t["duration_ms"])
        cell(ws, row, 7, t.get("distinct_values", ""))
        fl = cell(ws, row, 8, flag)
        if t["status"] != "ok":
            st.fill, st.font = BAD_FILL, BAD_FONT
        if ff:
            fl.fill, fl.font = WARN_FILL, WARN_FONT
        row += 1
    cell(ws, row, 1, "TOTAL", bold=True)
    cell(ws, row, 6, f"=SUM(F2:F{row - 1})", bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{row - 1}"
    auto_width(ws)

    # ── By Measure ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("By Measure")
    cols = ["Measure", "Tests", "OK", "Total (ms)", "Avg (ms)", "Max (ms)", "Min (ms)"]
    for c, h in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=h)
    header(ws, 1, len(cols))
    row = 2
    for a in aggregate_by_measure(timing_rows):
        cell(ws, row, 1, a["measure"])
        cell(ws, row, 2, a["tests"])
        cell(ws, row, 3, a["ok"])
        cell(ws, row, 4, a["total_ms"])
        cell(ws, row, 5, a["avg_ms"])
        cell(ws, row, 6, a["max_ms"])
        cell(ws, row, 7, a["min_ms"])
        row += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{row - 1}"
    auto_width(ws)

    # ── Slowest ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Slowest")
    cols = ["Rank", "Test ID", "Measure", "Context", "Duration (ms)", "Rows", "Distinct"]
    for c, h in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=h)
    header(ws, 1, len(cols))
    row = 2
    for rank, t in enumerate(slowest(timing_rows), 1):
        cell(ws, row, 1, rank)
        cell(ws, row, 2, t["test_id"])
        cell(ws, row, 3, t["measure"])
        cell(ws, row, 4, t["context"])
        cell(ws, row, 5, t["duration_ms"])
        cell(ws, row, 6, t["row_count"])
        cell(ws, row, 7, t.get("distinct_values", ""))
        row += 1
    ws.freeze_panes = "A2"
    auto_width(ws)

    # ── False-Fast Warnings ──────────────────────────────────────────────────
    ws = wb.create_sheet("False-Fast Warnings")
    if not flagged:
        ws["A1"] = "No false-fast warnings — every dimension filtered its measure."
        ws["A1"].font = OK_FONT
    else:
        ws["A1"] = ("distinct_values=1 with row_count>1 — the dimension may not "
                    "filter this measure (missing relationship path).")
        ws["A1"].font = Font(name="Arial", size=11, italic=True, color="9C6500")
        cols = ["Test ID", "Measure", "Context", "Duration (ms)", "Rows", "Distinct"]
        for c, h in enumerate(cols, 1):
            ws.cell(row=2, column=c, value=h)
        header(ws, 2, len(cols))
        row = 3
        for t in flagged:
            cell(ws, row, 1, t["test_id"])
            cell(ws, row, 2, t["measure"])
            cell(ws, row, 3, t["context"])
            cell(ws, row, 4, t["duration_ms"])
            cell(ws, row, 5, t["row_count"])
            cell(ws, row, 6, t.get("distinct_values", ""))
            row += 1
        ws.freeze_panes = "A3"
        auto_width(ws)

    wb.save(str(path))
