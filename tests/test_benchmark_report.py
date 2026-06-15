import pytest

from pbi_capture import benchmark_report as br
from pbi_capture.config import BenchmarkConfig

TIMING = [
    {"test_id": "b0001", "measure": "M One", "context": "grand_total",
     "status": "ok", "row_count": 1, "duration_ms": 10, "distinct_values": 1},
    {"test_id": "b0002", "measure": "M One", "context": "by_year",
     "status": "ok", "row_count": 5, "duration_ms": 200, "distinct_values": 1},  # false-fast
    {"test_id": "b0003", "measure": "M Two", "context": "by_year",
     "status": "ok", "row_count": 3, "duration_ms": 50, "distinct_values": 3},
    {"test_id": "b0004", "measure": "M Two", "context": "grand_total",
     "status": "error", "row_count": 0, "duration_ms": 5, "distinct_values": 0},
]


def test_aggregate_by_measure():
    agg = br.aggregate_by_measure(TIMING)
    assert [a["measure"] for a in agg] == ["M One", "M Two"]  # 210ms total > 50ms
    m1, m2 = agg
    assert (m1["tests"], m1["ok"], m1["total_ms"]) == (2, 2, 210)
    assert (m1["max_ms"], m1["min_ms"], m1["avg_ms"]) == (200, 10, 105)
    assert (m2["tests"], m2["ok"], m2["total_ms"]) == (2, 1, 50)  # error excluded from ok
    assert m2["avg_ms"] == 50


def test_false_fast():
    assert [t["test_id"] for t in br.false_fast(TIMING)] == ["b0002"]


def test_slowest_excludes_non_ok():
    assert [t["test_id"] for t in br.slowest(TIMING, top_n=2)] == ["b0002", "b0003"]


def test_write_report_creates_styled_workbook(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    cfg = BenchmarkConfig(label="b", measures=["M One", "M Two"],
                          single_slice_dimensions={"by_year": "'D'[Y]"})
    path = tmp_path / "b-report.xlsx"
    br.write_benchmark_report(
        path, cfg=cfg,
        counts={"ok": 3, "error": 1, "timeout": 0, "skipped": 0, "aborted_memory": 0},
        total=4, total_ms=265, timing_rows=TIMING, skipped=set(),
        smoke_results={}, memory_aborted=False)
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Summary", "All Tests", "By Measure",
                             "Slowest", "False-Fast Warnings"]


def test_write_report_handles_empty_and_clean(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    cfg = BenchmarkConfig(label="clean", measures=["M"])
    clean = [{"test_id": "b0001", "measure": "M", "context": "grand_total",
              "status": "ok", "row_count": 1, "duration_ms": 8, "distinct_values": 1}]
    path = tmp_path / "clean-report.xlsx"
    br.write_benchmark_report(
        path, cfg=cfg,
        counts={"ok": 1, "error": 0, "timeout": 0, "skipped": 0, "aborted_memory": 0},
        total=1, total_ms=8, timing_rows=clean, skipped=set(),
        smoke_results={}, memory_aborted=False)
    wb = openpyxl.load_workbook(path)
    # No false-fast rows → friendly note rather than a header row.
    assert "No false-fast" in str(wb["False-Fast Warnings"]["A1"].value)
